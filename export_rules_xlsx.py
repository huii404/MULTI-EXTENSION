# -*- coding: utf-8 -*-
import openpyxl
import os
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_all():
    font_family = "Times New Roman"
    
    # ----------------------------------------------------
    # DATA DEFINITIONS
    # ----------------------------------------------------
    
    # Overview / Architecture info
    overview_headers = ["Tầng xử lý (Tier)", "Công nghệ / Model AI", "Vai trò hệ thống", "Cơ chế hoạt động & Xử lý ngoại lệ"]
    overview_data = [
        ["Tầng 1 (Ưu tiên)", "Google Gemini 2.5 Flash", "Phân tích ngữ nghĩa sâu chính", "Phân tích sâu ngữ nghĩa, cảm xúc ẩn ý, ngữ cảnh phức tạp. Tự động luân chuyển (Key Rotation) khi gặp lỗi."],
        ["Tầng 2 (Dự phòng 1)", "Groq Llama 3.3 70B Versatile", "Dự phòng AI tốc độ cao", "Kích hoạt khi Gemini hết hạn mức (Quota 429). Sử dụng HTTP Header User-Agent chuẩn vượt qua Cloudflare."],
        ["Tầng 3 (Dự phòng 2)", "Rule-based Engine (Python)", "Bảo vệ hệ thống & Chạy Offline", "Kích hoạt khi cả 2 AI đều lỗi. Phân tích từ khóa, ngập ngừng, lặp từ, phủ định. Đảm bảo ứng dụng 100% không crash."]
    ]
    
    # Sheet 1: Main Classification Rules
    sheet1_headers = [
        "Cảm xúc chủ đạo", 
        "Điểm số từ khóa trực tiếp (Keywords)", 
        "Biểu hiện gián tiếp & Mẫu câu (Indirect & Sentence Patterns)", 
        "Tác động của Ngập ngừng, Lặp từ & Tiếng cười (Voice/Text Signals)", 
        "Điều kiện phân loại (Conditions)"
    ]
    
    sheet1_data = [
        [
            "STRESS\n(Căng thẳng / Bực bội)", 
            "Tức giận (+3): bực, tức, cáu, điên, phát điên, ức chế...\nÁp lực (+3): áp lực, stress, căng thẳng.\nKhó chịu (+2): khó chịu, phiền phức, ngán ngẩm...\nMệt mỏi (+2): mệt mỏi, mệt quá.", 
            "Ý kiến gián tiếp (+2):\n- 'sao cứ như vậy hoài vậy'\n- 'làm gì lâu vậy'\n- 'chịu không nổi'\n- 'tôi mệt với chuyện này rồi'\n\nCảm thán tiêu cực: Câu chứa cảm xúc tiêu cực có dấu '!' được cộng thêm Stress (+2).", 
            "Lặp từ nhấn mạnh: Lặp từ bực tức (ví dụ: 'không! không! không!') làm tăng điểm Stress.\n\n(Lưu ý: Các từ tiếng cười như 'kkkk' được loại trừ khỏi bộ lặp từ).", 
            "Điểm Stress là cao nhất trong 4 nhóm cảm xúc và >= 0.5."
        ],
        [
            "ANXIETY\n(Lo lắng / Hồi hộp)", 
            "Lo lắng (+3): lo, lo lắng, hồi hộp, bồn chồn, run...\nSợ hãi (+3): sợ, e là, sợ quá, tôi sợ, tôi lo...\nMơ hồ (+2): không biết, không chắc, hình như, có lẽ, có vẻ, liệu, nhỡ, lỡ...", 
            "Mẫu câu bất an (+2):\n- Câu hỏi chứa 'nếu... thì sao'\n- 'nhỡ...'\n- 'lỡ...'\n- 'mình phải làm sao'", 
            "Ngập ngừng (+2): Có dấu '...' hoặc từ đệm (ừm, ờ, à, hmm...)\n\nLặp từ hồi hộp (+2): Lặp từ đại từ ('tôi... tôi') hoặc lặp từ khác.", 
            "Điểm Anxiety là cao nhất và >= 0.5.\n(Lưu ý: Nếu chỉ có ngập ngừng mà không có từ khóa lo lắng nào khác thì điểm Anxiety sẽ bị loại bỏ để tránh nhầm lẫn)."
        ],
        [
            "SADNESS\n(Buồn bã)", 
            "Buồn (+3): buồn, buồn quá, đau lòng, tổn thương, không vui...\nThất vọng (+3): thất vọng, chán, chán nản, nản lòng, thất bại, tuyệt vọng, vô vọng...\nCô đơn (+3): cô đơn, một mình, nhớ...", 
            "Biểu hiện gián tiếp (+2):\n- 'dạo này tôi chẳng muốn làm gì cả'\n- 'mọi thứ hình như chẳng còn ý nghĩa'\n- 'hôm nay thật sự là một ngày tệ'\n\nPhủ định tích cực (+2): Cấu trúc dạng không + từ vui vẻ (ví dụ: 'không vui', 'không thích').", 
            "Góp phần tăng nhẹ cảm giác buồn bã gián tiếp khi kết hợp với ngập ngừng chậm rãi.", 
            "Điểm Sadness là cao nhất trong 4 nhóm và >= 0.5."
        ],
        [
            "HAPPY\n(Vui vẻ)", 
            "Hạnh phúc (+3): vui, vui vẻ, hạnh phúc, tuyệt vời, thích, thích quá, may quá, tốt quá...\nHào hứng (+3): hào hứng, phấn khích, thành công, chiến thắng...", 
            "Biểu hiện gián tiếp (+2):\n- 'cuối cùng cũng làm xong rồi'\n- 'tôi qua môn rồi'\n- 'hôm nay thật tuyệt'\n\nCảm thán tích cực (+2): Câu chứa cảm xúc vui vẻ có dấu '!'.", 
            "Tiếng cười chat lóng (+2): Các từ haha, hehe, hihi, hahaha, kkkk, kaka, kkk, kk.\n\nAI & Rule-based nhận diện chuẩn xác 'kkkk' là HAPPY thay vì lặp từ.", 
            "Điểm Happy là cao nhất trong 4 nhóm và >= 0.5."
        ],
        [
            "NEUTRAL\n(Bình thường)", 
            "Không có từ khóa cảm xúc nào hoặc điểm số của các từ khóa bị triệt tiêu bởi phủ định (ví dụ: 'tôi không lo', 'tôi không buồn').", 
            "Các câu giao tiếp xã giao thông thường (ví dụ: 'xin chào', 'bạn khỏe không', 'cảm ơn bạn').", 
            "Các từ ngập ngừng đệm thuần túy không mang tính lo âu (ví dụ: 'ừm... để tôi xem').", 
            "Tất cả điểm số cảm xúc đều < 0.5."
        ]
    ]
    
    # Sheet 2: Detailed Weights Matrix
    sheet2_headers = [
        "Đặc trưng / Tín hiệu phát hiện",
        "STRESS",
        "ANXIETY",
        "SADNESS",
        "HAPPY",
        "Mô tả chi tiết từ khóa / Ví dụ thực tế"
    ]
    
    sheet2_data = [
        ["Từ khóa Tức giận", "+3", "0", "0", "0", "bực, tức, cáu, điên, phát điên, ức chế..."],
        ["Từ khóa Khó chịu", "+2", "0", "+1", "0", "khó chịu, phiền, phiền phức, ngán, ngán ngẩm..."],
        ["Từ khóa Áp lực", "+3", "+2", "0", "0", "áp lực, stress, căng thẳng..."],
        ["Từ khóa Lo lắng", "0", "+3", "0", "0", "lo, lo lắng, hồi hộp, bất an, bồn chồn, run..."],
        ["Từ khóa Sợ hãi", "0", "+3", "+1", "0", "sợ, e là, sợ quá, sợ rằng..."],
        ["Ngập ngừng (Do từ đệm hoặc dấu ...)", "0", "+2", "0", "0", "Có dấu '...' hoặc từ đệm ngập ngừng: ừm, ờ, à, hmm, um..."],
        ["Lặp từ do hồi hộp", "0", "+2", "0", "0", "Lặp từ đứng cạnh nhau (ví dụ: 'sợ... sợ', 'không... không')"],
        ["Lặp đại từ đặc biệt 'tôi... tôi'", "0", "+3", "0", "0", "Lặp đại từ nhân xưng 'tôi... tôi' (cấp độ ngập ngừng cao)"],
        ["Từ khóa Không chắc chắn", "0", "+2", "0", "0", "hình như, có lẽ, có vẻ, liệu, nhỡ, lỡ, hay là, có khi..."],
        ["Từ khóa Thất vọng", "0", "+1", "+3", "0", "thất vọng, chán, chán nản, nản lòng, thất bại, tuyệt vọng, vô vọng..."],
        ["Từ khóa Cô đơn", "0", "+1", "+3", "0", "cô đơn, một mình, nhớ, nhớ ai đó..."],
        ["Từ khóa Buồn bã", "0", "0", "+3", "0", "buồn, buồn quá, đau lòng, tổn thương, không vui..."],
        ["Từ khóa Hạnh phúc / Vui vẻ", "0", "0", "0", "+3", "vui, vui vẻ, hạnh phúc, tuyệt vời, thích, thích quá..."],
        ["Từ khóa Hào hứng", "0", "0", "0", "+3", "hào hứng, phấn khích, thành công, chiến thắng..."],
        ["Tiếng cười & Từ lóng ('kkkk', 'haha')", "0", "0", "0", "+2", "haha, hahaha, hehe, hihi, kkkk, kaka, kkk, kk..."],
        ["Không + Từ tích cực (Phủ định)", "0", "0", "+2", "-2", "Khi phát hiện phủ định từ vui vẻ (ví dụ: 'không vui', 'chẳng thích')"],
        ["Không + Từ tiêu cực (Phủ định)", "-2", "-2", "-2", "0", "Giảm điểm của cảm xúc tiêu cực tương ứng (ví dụ: 'không lo', 'không sợ')"],
        ["Câu hỏi bất an", "0", "+2", "0", "0", "Câu hỏi (?) chứa từ nghi vấn lo lắng: 'liệu', 'nếu... thì sao', 'phải làm sao'"],
        ["Câu cảm thán tích cực", "0", "0", "0", "+2", "Câu cảm thán (!) khi câu nói có chứa từ khóa vui vẻ (HAPPY)"],
        ["Câu cảm thán tiêu cực", "+2", "+1", "+1", "0", "Câu cảm thán (!) khi câu nói chứa các từ khóa tiêu cực (Stress/Anxiety/Sadness)"]
    ]
    
    # ----------------------------------------------------
    # 1. EXPORT TO CSV (Bang_Quy_tac_Phan_loai_Cam_xuc.csv)
    # ----------------------------------------------------
    csv_file = "Bang_Quy_tac_Phan_loai_Cam_xuc.csv"
    with open(csv_file, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        
        # Section 1: Overview
        writer.writerow(["=== KIẾN TRÚC HỆ THỐNG PHÂN TÍCH CẢM XÚC (3-TIER HYBRID AI) ==="])
        writer.writerow(overview_headers)
        writer.writerows(overview_data)
        writer.writerow([])
        
        # Section 2: Main Classification Rules
        writer.writerow(["=== BẢNG QUY TẮC PHÂN LOẠI CẢM XÚC CHỦ ĐẠO ==="])
        writer.writerow(sheet1_headers)
        writer.writerows(sheet1_data)
        writer.writerow([])
        
        # Section 3: Scoring Matrix
        writer.writerow(["=== BẢNG HỆ SỐ TRỌNG SỐ ĐIỂM CHI TIẾT (SCORING MATRIX) ==="])
        writer.writerow(sheet2_headers)
        writer.writerows(sheet2_data)

    print(f"Exported CSV successfully to {os.path.abspath(csv_file)}")

    # ----------------------------------------------------
    # 2. EXPORT TO STYLED EXCEL (Bang_Quy_tac_Phan_loai_Cam_xuc.xlsx)
    # ----------------------------------------------------
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # --- SHEET 1: Quy tắc cảm xúc ---
    ws1 = wb.active
    ws1.title = "Quy tắc cảm xúc & Kiến trúc"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A2:E2")
    title1 = ws1["A2"]
    title1.value = "BẢNG QUY TẮC PHÂN LOẠI CẢM XÚC - EMOTION ANALYZER (HYBRID AI)"
    title1.font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    title1.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 40
    
    # Overview Table (Row 4 to 8)
    ws1.merge_cells("A4:D4")
    ws1["A4"].value = "KIẾN TRÚC HỆ THỐNG LAI 3 TẦNG (RESILIENCE PIPELINE)"
    ws1["A4"].font = Font(name=font_family, size=12, bold=True, color="1F4E78")
    
    for c_i, h in enumerate(overview_headers, 1):
        cell = ws1.cell(row=5, column=c_i)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border
    ws1.row_dimensions[5].height = 25
    
    for r_i, r_data in enumerate(overview_data, 6):
        for c_i, val in enumerate(r_data, 1):
            cell = ws1.cell(row=r_i, column=c_i)
            cell.value = val
            cell.font = Font(name=font_family, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = cell_border
        ws1.row_dimensions[r_i].height = 35
        
    # Main Rules Table (Row 11 onwards)
    ws1.merge_cells("A10:E10")
    ws1["A10"].value = "BẢNG QUY TẮC PHÂN LOẠI CẢM XÚC CHỦ ĐẠO"
    ws1["A10"].font = Font(name=font_family, size=12, bold=True, color="1F4E78")
    
    for c_i, h in enumerate(sheet1_headers, 1):
        cell = ws1.cell(row=11, column=c_i)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border
    ws1.row_dimensions[11].height = 30
    
    emotion_styles = {
        "STRESS": {"fill": "FADBD8", "font_color": "78281F"},
        "ANXIETY": {"fill": "FCF3CF", "font_color": "7E5109"},
        "SADNESS": {"fill": "D6EAF8", "font_color": "1B4F72"},
        "HAPPY": {"fill": "D4EFDF", "font_color": "145A32"},
        "NEUTRAL": {"fill": "EAECEE", "font_color": "5D6D7E"}
    }
    
    for r_i, row_data in enumerate(sheet1_data, 12):
        emo_key = "NEUTRAL"
        for key in emotion_styles:
            if key in row_data[0]:
                emo_key = key
                break
        style = emotion_styles[emo_key]
        
        cell_a = ws1.cell(row=r_i, column=1)
        cell_a.value = row_data[0]
        cell_a.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        cell_a.font = Font(name=font_family, size=11, bold=True, color=style["font_color"])
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a.border = cell_border
        
        bg_color = "F9FBFD" if r_i % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        for c_i in range(2, 6):
            cell = ws1.cell(row=r_i, column=c_i)
            cell.value = row_data[c_i-1]
            cell.fill = row_fill
            cell.font = Font(name=font_family, size=10, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = cell_border
            
        ws1.row_dimensions[r_i].height = 145
        
    column_widths = {"A": 22, "B": 42, "C": 45, "D": 42, "E": 38}
    for col_letter, width in column_widths.items():
        ws1.column_dimensions[col_letter].width = width
        
    # --- SHEET 2: Trọng số điểm chi tiết ---
    ws2 = wb.create_sheet(title="Trọng số điểm chi tiết")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A2:F2")
    title2 = ws2["A2"]
    title2.value = "BẢNG HỆ SỐ TRỌNG SỐ ĐIỂM CHI TIẾT (SCORING MATRIX)"
    title2.font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    title2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 40
    
    for col_idx, header in enumerate(sheet2_headers, 1):
        cell = ws2.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border
    ws2.row_dimensions[4].height = 30
    
    col_fills = {
        2: PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        3: PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        4: PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
        5: PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
    }
    
    bold_times = Font(name=font_family, size=10, bold=True)
    normal_times = Font(name=font_family, size=10)
    gray_times = Font(name=font_family, size=10, color="888888")
    red_bold_times = Font(name=font_family, size=10, bold=True, color="FF0000")
    
    for r_idx, row_data in enumerate(sheet2_data, 5):
        cell_1 = ws2.cell(row=r_idx, column=1)
        cell_1.value = row_data[0]
        cell_1.font = bold_times
        cell_1.alignment = Alignment(horizontal="left", vertical="center")
        cell_1.border = cell_border
        
        for c_idx in range(2, 6):
            cell = ws2.cell(row=r_idx, column=c_idx)
            val_str = row_data[c_idx-1]
            cell.value = val_str
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if val_str == "0":
                cell.font = gray_times
            elif val_str.startswith("-"):
                cell.font = red_bold_times
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            else:
                cell.font = Font(name=font_family, size=10, bold=True, color="000000")
                cell.fill = col_fills[c_idx]
                
        cell_6 = ws2.cell(row=r_idx, column=6)
        cell_6.value = row_data[5]
        cell_6.font = normal_times
        cell_6.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_6.border = cell_border
        
        ws2.row_dimensions[r_idx].height = 25
        
    ws2_widths = {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 65}
    for col_letter, width in ws2_widths.items():
        ws2.column_dimensions[col_letter].width = width
        
    excel_file = "Bang_Quy_tac_Phan_loai_Cam_xuc.xlsx"
    try:
        wb.save(excel_file)
    except PermissionError:
        excel_file = "Bang_Quy_tac_Phan_loai_Cam_xuc_v2.xlsx"
        wb.save(excel_file)
        
    print(f"Exported Excel successfully to {os.path.abspath(excel_file)}")

if __name__ == "__main__":
    export_all()
